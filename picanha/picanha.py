from flask import Flask, render_template, request
from flask_wtf.csrf import CSRFProtect
from openpyxl import load_workbook


from api_access import acceso


import os

app = Flask(__name__)
app.secret_key = 'MY_SECRET_KEY_PORRA!'
csrf = CSRFProtect()
csrf.init_app(app)

def find_files(dbx):
    value = request.form.getlist('check')
    names = []
    mappings = []
    blocklists = []
    for ruta in value:
        for auxiliar in dbx.files_list_folder(ruta).entries:
            names.append(auxiliar.name)
            if auxiliar.name.find("apping") > 0 and auxiliar.name.find("xlsx") > -1:
                mappings.append(auxiliar.path_display)
            elif auxiliar.name.find("locklist") > 0:
                blocklists.append(auxiliar.path_display)
    os.chdir(r"C:\Users\Michele\python\picanha\mappings_downloads")
    for name, route in zip(names, mappings):
        dbx.files_download_to_file(name, route)
def getNameOfWebsites():
    #change the following line before operate in another system
    files = os.listdir(r"C:\Users\Michele\python\picanha\mappings_downloads")
    namefiles = []
    for basura in files:
        basura = basura.lower()
        basura = basura.replace('yieldlove', '')
        basura = basura.replace('mapping', '')
        basura = basura.replace("_", "")
        basura = basura.replace(".xlsx","")
        if "blacklist" in basura:
            basura = basura.replace("blacklist","")
        namefiles.append(basura)
    return namefiles, files

def GetDataOfExcel(routs):
    mapping_folder = os.getcwd()+"\\"
    listac = [] #this will contain the result of everything
    for file_route in routs:
        complete_route = mapping_folder + file_route
        #in the following lines we get the file, the total amout of rows (rows) and we don't need the amount of columns
        wb = load_workbook(complete_route)
        ws = wb.active
        rows = ws.max_row
        #now we are going to read all the data from the file based in the total line of the rows
        #we will get two different lists (listaa and listab) and finally merge it in listaa
        listaa = []
        listab = []
        for cell in range(2, rows):
            if ws['A' + str(cell)].value == None:
                break
            else:
                listaa.append(ws['A' + str(cell)].value)
                listab.append(ws['B' + str(cell)].value)
        listaa.append("-")
        listab.append("-")
        listac = zip(listaa,listab)
        return listac

@app.route('/', methods = ['GET', 'POST'])
def mapping():
    dbx = acceso.dbx
    NameOfTheSites = []
    RouteFromFiles = []
    listac = []
    basura = dbx.files_list_folder('/AAA_Webseiten').entries
    if request.method == 'POST':
        find_files(dbx)
        NameOfTheSites, RoutFromFiles = getNameOfWebsites()
        listac = GetDataOfExcel(RoutFromFiles)

        return render_template("table.html", mapping =  listac)

    return render_template("mapping.html", folders = basura)



@app.route('/backup', methods = ['GET', 'POST'])
def backup():
    dbx = acceso.dbx
    basura = dbx.files_list_folder('/AAA_Webseiten').entries
    if request.method == 'POST':
        value = request.form.getlist('check')
        names = []
        mappings = []
        blocklists = []
        for ruta in value:
            for auxiliar in dbx.files_list_folder(ruta).entries:
                names.append(auxiliar.name)
                if auxiliar.name.find("apping") > 0:
                    mappings.append(auxiliar.path_display)
                elif auxiliar.name.find("locklist") > 0:
                    blocklists.append(auxiliar.path_display)
        os.chdir(r"C:\Users\Coult\python\udemy\cursopagado\picanha\mappings_downloads")
        for name, route in zip(names, mappings):
            dbx.files_download_to_file(name, route)
        return "deveria ter descarregado o arquivo"
    return render_template("mapping.html", folders = basura)




if __name__ == '__main__':
    app.run(debug = True, port = 8000)
